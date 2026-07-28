"""Paywall enforcement in the review path + free-tier export watermark
(R-004, TS-087/TS-088). Before this, authorize_review's logic was correct but
had exactly one caller — its own HTTP endpoint — so any client could run
unlimited reviews by simply not calling it, and Grant.watermark was computed
and never consumed anywhere.
"""

from __future__ import annotations

import hashlib
import hmac
import json
import uuid

import pytest
from fastapi.testclient import TestClient

import app.modules.auth.models  # noqa: F401
import app.modules.billing.models  # noqa: F401
import app.modules.drafting.models  # noqa: F401
import app.modules.findings.models  # noqa: F401
import app.modules.ingestion.models  # noqa: F401
import app.modules.review.models  # noqa: F401
from app.core.config import Settings
from app.core.db import Base
from app.main import create_app
from app.modules.billing.providers.base import OrderHandle
from app.modules.billing.providers.razorpay import RazorpayProvider

SECRET = "dev-razorpay-secret"

_FULL_MODULES = (
    "health,rulepacks,auth,billing,ingestion,risk,findings,review,drafting,export"
)


def _make_client(**settings_kwargs) -> TestClient:
    app = create_app(
        Settings(
            enabled_modules=settings_kwargs.pop("enabled_modules", _FULL_MODULES),
            database_url="sqlite:///:memory:",
            razorpay_key_id="rzp_test_fake",
            razorpay_key_secret="fake_secret",
            razorpay_webhook_secret=SECRET,
            **settings_kwargs,
        )
    )
    Base.metadata.create_all(app.state.ctx.registry.require("db.engine"))
    return TestClient(app)


@pytest.fixture
def client(monkeypatch) -> TestClient:
    def fake_create_order(self, req):
        return OrderHandle(
            provider="razorpay",
            order_id=f"order_{req.idempotency_key[:16]}",
            amount_minor=req.amount_minor,
            currency=req.currency,
            checkout_payload={"key": self._key_id, "amount": req.amount_minor},
        )

    monkeypatch.setattr(RazorpayProvider, "create_order", fake_create_order)
    return _make_client()


def _auth(client, email="paywall@x.com"):
    client.post(
        "/api/auth/signup",
        json={"email": email, "password": "hunter2hunter2", "workspace_name": "Acme"},
    )
    r = client.post("/api/auth/login", json={"email": email, "password": "hunter2hunter2"})
    return {"authorization": f"Bearer {r.json()['access_token']}"}, r.json()["workspace_id"]


def _create_opportunity(client, headers, title="Bridge"):
    return client.post(
        "/api/ingestion/opportunities", json={"title": title}, headers=headers
    ).json()["id"]


def _sign(body: dict):
    raw = json.dumps(body).encode()
    sig = hmac.new(SECRET.encode(), raw, hashlib.sha256).hexdigest()
    return raw, sig


def _set_plan_via_webhook(client, headers, workspace_id, plan="pro"):
    """Real checkout -> real webhook, resolving the grant from the intent
    (R-005) rather than the pre-R-005 shape that trusted notes.workspace_id/
    notes.plan directly."""
    checkout = client.post(
        "/api/billing/checkout",
        json={"kind": "subscription", "plan": plan},
        headers=headers,
    )
    assert checkout.status_code == 200, checkout.text
    intent = checkout.json()

    body = {
        "id": f"evt_activate_{workspace_id}",
        "event": "subscription.activated",
        "payload": {
            "payment": {
                "entity": {
                    "amount": intent["amount_minor"],
                    "notes": {"intent_id": intent["intent_id"]},
                }
            }
        },
    }
    raw, sig = _sign(body)
    r = client.post(
        "/api/billing/webhooks/razorpay", content=raw, headers={"X-Razorpay-Signature": sig}
    )
    assert r.status_code == 200, r.text


# ---- A1: free workspace, second opportunity is paywalled -------------------


def test_second_review_on_free_workspace_is_paywalled(client):
    headers, _ = _auth(client)
    opp1 = _create_opportunity(client, headers, "First")
    r1 = client.post(f"/api/risk/opportunities/{opp1}/run", headers=headers)
    assert r1.status_code == 200, r1.text
    assert r1.json()["watermark"] is True

    opp2 = _create_opportunity(client, headers, "Second")
    r2 = client.post(f"/api/risk/opportunities/{opp2}/run", headers=headers)
    assert r2.status_code == 402
    assert r2.json()["detail"]["code"] == "free_exhausted"
    assert r2.json()["detail"]["upsell"]["paygo_price_inr_paise"] == 750_000

    # nothing was processed for the blocked opportunity
    findings = client.get(f"/api/findings/opportunities/{opp2}", headers=headers).json()
    assert findings["findings"] == []


# ---- A2: pro workspace at monthly quota -------------------------------------


def test_pro_workspace_quota_exhausted(client):
    headers, workspace_id = _auth(client, "proquota@x.com")
    _set_plan_via_webhook(client, headers, workspace_id, "pro")

    for i in range(10):
        opp = _create_opportunity(client, headers, f"Opp {i}")
        r = client.post(f"/api/risk/opportunities/{opp}/run", headers=headers)
        assert r.status_code == 200, r.text
        assert r.json()["watermark"] is False

    opp_11 = _create_opportunity(client, headers, "Opp 11")
    r = client.post(f"/api/risk/opportunities/{opp_11}/run", headers=headers)
    assert r.status_code == 402
    assert r.json()["detail"]["code"] == "quota_exhausted"


# ---- paygo: requires_payment was computed but never enforced (R-008/TS-091) -


def _set_plan_direct(client, workspace_id, plan):
    """Test-only seed, bypassing checkout/webhook — mirrors the same
    WorkspaceAdmin capability the webhook itself calls on payment success."""
    reg = client.app.state.ctx.registry
    factory = reg.require("db.sessionmaker")
    workspace_factory = reg.require("auth.workspace_factory")
    with factory() as s:
        workspace_factory(s).set_plan(workspace_id, plan)
        s.commit()


def test_paygo_workspace_blocked_until_its_own_opportunity_is_paid(client):
    """Regression test: plans.authorize()'s Grant(requires_payment=True) for
    a paygo workspace was, until this fix, advisory only — nothing checked
    it, so a paygo workspace could run unlimited unpaid reviews. Found while
    wiring the checkout UI (R-008/TS-091)."""
    headers, workspace_id = _auth(client, "paygoenforce@x.com")
    _set_plan_direct(client, workspace_id, "paygo")

    opp = _create_opportunity(client, headers, "Unpaid")
    r = client.post(f"/api/risk/opportunities/{opp}/run", headers=headers)
    assert r.status_code == 402
    assert r.json()["detail"]["code"] == "paygo_payment_required"
    assert r.json()["detail"]["upsell"]["paygo_price_inr_paise"] == 750_000

    # nothing was processed for the blocked opportunity
    findings = client.get(f"/api/findings/opportunities/{opp}", headers=headers).json()
    assert findings["findings"] == []

    checkout = client.post(
        "/api/billing/checkout",
        json={"kind": "paygo", "opportunity_id": opp},
        headers=headers,
    )
    assert checkout.status_code == 200, checkout.text
    intent = checkout.json()
    body = {
        "id": f"evt_paygo_{opp}",
        "event": "order.paid",
        "payload": {
            "payment": {
                "entity": {
                    "amount": intent["amount_minor"],
                    "notes": {"intent_id": intent["intent_id"]},
                }
            }
        },
    }
    raw, sig = _sign(body)
    wh = client.post(
        "/api/billing/webhooks/razorpay", content=raw, headers={"X-Razorpay-Signature": sig}
    )
    assert wh.status_code == 200, wh.text

    r2 = client.post(f"/api/risk/opportunities/{opp}/run", headers=headers)
    assert r2.status_code == 200, r2.text
    assert r2.json()["watermark"] is False

    # paying for THIS opportunity does not unlock a DIFFERENT one
    opp_other = _create_opportunity(client, headers, "Still unpaid")
    r3 = client.post(f"/api/risk/opportunities/{opp_other}/run", headers=headers)
    assert r3.status_code == 402
    assert r3.json()["detail"]["code"] == "paygo_payment_required"


def test_free_exhausted_workspace_can_pay_per_tender_from_the_paywall(client):
    """The free_exhausted upsell carries the blocked opportunity's id
    (plans.authorize()'s opportunity_id param, R-008/TS-091) so the paywall
    can check out a paygo payment for THAT tender directly. Paying elects
    the workspace into the paygo plan (service.py's `_on_payment_succeeded`
    sets plan=intent.plan for every kind, not just subscriptions) and
    unlocks exactly the paid-for opportunity."""
    headers, _ = _auth(client, "freetopaygo@x.com")
    opp1 = _create_opportunity(client, headers, "Free one")
    assert client.post(f"/api/risk/opportunities/{opp1}/run", headers=headers).status_code == 200
    opp = _create_opportunity(client, headers, "Blocked")
    blocked = client.post(f"/api/risk/opportunities/{opp}/run", headers=headers)
    assert blocked.status_code == 402
    assert blocked.json()["detail"]["code"] == "free_exhausted"
    assert blocked.json()["detail"]["upsell"]["opportunity_id"] == opp

    checkout = client.post(
        "/api/billing/checkout",
        json={"kind": "paygo", "opportunity_id": opp},
        headers=headers,
    )
    assert checkout.status_code == 200, checkout.text
    intent = checkout.json()
    body = {
        "id": f"evt_free_to_paygo_{opp}",
        "event": "order.paid",
        "payload": {
            "payment": {
                "entity": {
                    "amount": intent["amount_minor"],
                    "notes": {"intent_id": intent["intent_id"]},
                }
            }
        },
    }
    raw, sig = _sign(body)
    wh = client.post(
        "/api/billing/webhooks/razorpay", content=raw, headers={"X-Razorpay-Signature": sig}
    )
    assert wh.status_code == 200, wh.text

    r = client.post(f"/api/risk/opportunities/{opp}/run", headers=headers)
    assert r.status_code == 200, r.text
    assert r.json()["watermark"] is False

    status = client.get("/api/billing/status", headers=headers).json()
    assert status["plan"] == "paygo"


# ---- A3: re-processing an already-metered opportunity is free --------------


def test_reprocessing_metered_opportunity_is_free(client):
    headers, workspace_id = _auth(client, "reprocess@x.com")
    opp = _create_opportunity(client, headers)
    r1 = client.post(f"/api/risk/opportunities/{opp}/run", headers=headers)
    assert r1.status_code == 200

    # re-run on the SAME opportunity (e.g. after an addendum) must not consume
    # the free review a second time
    r2 = client.post(f"/api/risk/opportunities/{opp}/run", headers=headers)
    assert r2.status_code == 200

    # confirm no second review_started usage event was written
    reg = client.app.state.ctx.registry
    factory = reg.require("db.sessionmaker")
    from sqlalchemy import select

    from app.modules.billing.models import UsageEvent

    with factory() as s:
        rows = s.scalars(
            select(UsageEvent).where(
                UsageEvent.workspace_id == uuid.UUID(workspace_id),
                UsageEvent.event == "review_started",
            )
        ).all()
        assert len(rows) == 1


# ---- A5: billing disabled degrades gracefully in dev, refuses in prod ------


def test_billing_disabled_review_proceeds_unmetered_in_dev():
    client = _make_client(enabled_modules="health,rulepacks,auth,ingestion,risk")
    headers, _ = _auth(client, "nobilling@x.com")
    opp1 = _create_opportunity(client, headers, "First")
    r1 = client.post(f"/api/risk/opportunities/{opp1}/run", headers=headers)
    assert r1.status_code == 200
    opp2 = _create_opportunity(client, headers, "Second")
    r2 = client.post(f"/api/risk/opportunities/{opp2}/run", headers=headers)
    assert r2.status_code == 200  # unmetered — billing module absent, dev posture


def test_billing_disabled_review_refuses_in_production():
    client = _make_client(
        enabled_modules="health,rulepacks,auth,ingestion,risk", env="production"
    )
    headers, _ = _auth(client, "prodnobilling@x.com")
    opp = _create_opportunity(client, headers)
    r = client.post(f"/api/risk/opportunities/{opp}/run", headers=headers)
    assert r.status_code == 503
    assert r.json()["detail"] == "billing_unavailable"


# ---- A8: paywall_hit fires once per 402 -------------------------------------


def test_paywall_hit_event_fires_on_block(client):
    headers, _ = _auth(client, "eventtest@x.com")
    hits = []
    client.app.state.ctx.events.subscribe(
        "billing.paywall_hit", lambda event, payload: hits.append(payload)
    )
    _create_opportunity(client, headers, "First")
    opp1 = _create_opportunity(client, headers, "First2")
    client.post(f"/api/risk/opportunities/{opp1}/run", headers=headers)
    opp2 = _create_opportunity(client, headers, "Second")
    r = client.post(f"/api/risk/opportunities/{opp2}/run", headers=headers)
    assert r.status_code == 402
    assert len(hits) == 1
    assert hits[0]["code"] == "free_exhausted"


# ---- A6/A7: export watermark ------------------------------------------------


def _prepare_reviewed_opportunity(client, headers, opp_id):
    """Run risk, accept every finding, so the export gate opens."""
    findings = client.post(
        f"/api/risk/opportunities/{opp_id}/run", headers=headers
    ).json()["findings"]
    all_findings = client.get(
        f"/api/findings/opportunities/{opp_id}", headers=headers
    ).json()["findings"]
    for f in all_findings:
        client.post(
            f"/api/review/findings/{f['id']}",
            json={"decision": "accepted"},
            headers=headers,
        )
    return findings


def _xlsx_header_texts(content: bytes) -> str:
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # row 2 is the stamp_line (see render.py); header/footer carry the mark too.
    return "\n".join(
        [
            str(ws.cell(row=2, column=1).value or ""),
            str(ws.oddHeader.center.text or ""),
            str(ws.oddFooter.center.text or ""),
        ]
    )


def test_free_workspace_export_is_watermarked(client):
    headers, _ = _auth(client, "freeexport@x.com")
    opp = _create_opportunity(client, headers)
    _prepare_reviewed_opportunity(client, headers, opp)
    r = client.get(f"/api/export/opportunities/{opp}?format=xlsx", headers=headers)
    assert r.status_code == 200, r.text
    assert "FREE REVIEW" in _xlsx_header_texts(r.content)


def test_paid_workspace_export_is_not_watermarked(client):
    headers, workspace_id = _auth(client, "paidexport@x.com")
    _set_plan_via_webhook(client, headers, workspace_id, "pro")
    opp = _create_opportunity(client, headers)
    _prepare_reviewed_opportunity(client, headers, opp)
    r = client.get(f"/api/export/opportunities/{opp}?format=xlsx", headers=headers)
    assert r.status_code == 200, r.text
    assert "FREE REVIEW" not in _xlsx_header_texts(r.content)


def test_free_and_paid_exports_have_identical_findings(client):
    """The watermark marks the document, never the content (R-004 §B.4)."""
    free_headers, _ = _auth(client, "contentfree@x.com")
    opp_free = _create_opportunity(client, free_headers, "Same Tender")
    findings_free = _prepare_reviewed_opportunity(client, free_headers, opp_free)

    paid_headers, paid_ws = _auth(client, "contentpaid@x.com")
    _set_plan_via_webhook(client, paid_headers, paid_ws, "pro")
    opp_paid = _create_opportunity(client, paid_headers, "Same Tender")
    findings_paid = _prepare_reviewed_opportunity(client, paid_headers, opp_paid)

    free_titles = sorted(f["title"] for f in findings_free)
    paid_titles = sorted(f["title"] for f in findings_paid)
    assert free_titles == paid_titles
