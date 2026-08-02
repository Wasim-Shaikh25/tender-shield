"""Text extraction from uploaded files (Doc §6.1). Digital PDFs via pypdf,
spreadsheets via openpyxl, DOCX via python-docx, CSV/text directly. Standalone
images (PNG/JPG/TIFF) are routed to an injected OCR provider. Page/paragraph
markers ([pN]) are emitted so downstream deadline/clause extraction can cite pages.
Scanned/image PDFs and images are flagged `needs_ocr` when no provider is configured
and degrade honestly (Doc §12.4)."""

from __future__ import annotations

import io
import zipfile

_MIN_PAGE_CHARS = 10  # a page with less digital text than this is "empty"


def extract_text(filename: str, data: bytes) -> str:
    """Simple digital-only extraction (no OCR). Kept for callers that pass text."""
    name = filename.lower()
    if name.endswith(".pdf"):
        return _pdf(data)
    if name.endswith((".xlsx", ".xlsm")):
        return _xlsx(data)
    if name.endswith(".csv"):
        return _csv(data)
    if name.endswith(".docx"):
        return _docx(data)
    if name.endswith(".zip"):
        return _zip(data, ocr=None)[0]
    if name.endswith((".png", ".jpg", ".jpeg", ".tiff", ".tif")):
        return ""
    if name.endswith((".txt", ".md")):
        return data.decode("utf-8", errors="replace")
    return data.decode("utf-8", errors="replace")


def _join_pages(pages: list[str]) -> str:
    return "\n".join(f"[p{i + 1}]\n{p}" for i, p in enumerate(pages))


def _pdf_pages(data: bytes) -> list[str]:
    import pypdf

    reader = pypdf.PdfReader(io.BytesIO(data))
    return [(page.extract_text() or "") for page in reader.pages]


def extract_upload(filename: str, data: bytes, ocr=None) -> tuple[str, str]:
    """Extraction for uploads. Returns (text, ocr_status) where ocr_status is
    one of: done | ocr_applied | needs_ocr. DOCX, XLSX, CSV, ZIP and plain text are
    extracted directly. Standalone images and PDFs with no digital text layer are
    OCR'd when a provider is given, else flagged needs_ocr."""
    name = filename.lower()
    if name.endswith((".png", ".jpg", ".jpeg", ".tiff", ".tif")):
        if ocr is not None and getattr(ocr, "name", "null") != "null" and hasattr(ocr, "ocr_image"):
            return _join_pages([ocr.ocr_image(data)]), "ocr_applied"
        return "", "needs_ocr"
    if name.endswith(".zip"):
        text, status = _zip(data, ocr=ocr)
        return text, status
    if not name.endswith(".pdf"):
        return extract_text(filename, data), "done"

    pages = _pdf_pages(data)
    has_text = any(len(p.strip()) >= _MIN_PAGE_CHARS for p in pages)
    if pages and not has_text:
        if ocr is not None and getattr(ocr, "name", "null") != "null":
            return _join_pages(ocr.ocr_pdf(data)), "ocr_applied"
        return _join_pages(pages), "needs_ocr"
    return _join_pages(pages), "done"


def looks_like_boq_csv(filename: str, data: bytes) -> bool:
    name = filename.lower()
    if not name.endswith(".csv"):
        return False
    header = data[:200].decode("utf-8", errors="replace").lower()
    return "description" in header and ("qty" in header or "rate" in header)


def _pdf(data: bytes) -> str:
    import pypdf

    reader = pypdf.PdfReader(io.BytesIO(data))
    return "\n".join(
        f"[p{i + 1}]\n{(page.extract_text() or '')}" for i, page in enumerate(reader.pages)
    )


def _xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        lines.append(f"[sheet:{ws.title}]")
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(f"[p{row_idx}]")
                lines.append("\t".join(cells))
    return "\n".join(lines)


def _csv(data: bytes) -> str:
    text = data.decode("utf-8", errors="replace")
    lines: list[str] = []
    for row_idx, line in enumerate(text.splitlines(), start=1):
        if line.strip():
            lines.append(f"[p{row_idx}]")
            lines.append(line)
    return "\n".join(lines)


def _docx(data: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(data))
    lines: list[str] = []
    para_idx = 0
    for block in doc.element.body:
        tag = block.tag.split("}")[-1] if "}" in block.tag else block.tag
        if tag == "p":
            para = doc.paragraphs[para_idx]
            text = para.text.strip()
            para_idx += 1
            if text:
                lines.append(f"[p{para_idx}]")
                lines.append(text)
        elif tag == "tbl":
            # Render table rows as tab-separated cells with a shared paragraph marker.
            table_text: list[str] = []
            for row in block.findall(".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tr"):
                cells = [cell.text.strip() for cell in row.findall(".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tc")]
                if any(cells):
                    table_text.append("\t".join(cells))
            if table_text:
                para_idx += 1
                lines.append(f"[p{para_idx}]")
                lines.extend(table_text)
    return "\n".join(lines)


def _zip(data: bytes, ocr=None) -> tuple[str, str]:
    """Extract all recognized files inside a ZIP archive. Each file is prefixed with
    `[file:<name>]` and its contained text is emitted as produced by `extract_upload`.
    Returns (combined_text, ocr_status) where status is the most degraded of all
    members: needs_ocr > ocr_applied > done."""
    statuses: set[str] = set()
    parts: list[str] = []
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            name = info.filename
            lower = name.lower()
            if lower.endswith(".zip"):
                continue  # no nested zip extraction to avoid recursion/explosion
            if not any(
                lower.endswith(ext)
                for ext in (
                    ".pdf", ".docx", ".xlsx", ".xlsm", ".csv", ".txt", ".md",
                    ".png", ".jpg", ".jpeg", ".tiff", ".tif",
                )
            ):
                continue
            file_data = zf.read(info.filename)
            if not file_data:
                continue
            text, status = extract_upload(name, file_data, ocr=ocr)
            if text:
                parts.append(f"[file:{name}]\n{text}")
            statuses.add(status)
    status = "done"
    if "ocr_applied" in statuses:
        status = "ocr_applied"
    if "needs_ocr" in statuses:
        status = "needs_ocr"
    return "\n\n".join(parts), status


def xlsx_to_rows(data: bytes) -> list[list[str]] | None:
    """First sheet → list of row lists, for BOQ canonical CSV conversion."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c) for c in row]
        if any(cells):
            rows.append(cells)
    return rows if rows else None
