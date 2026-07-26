"""Export: Bid Review Pack download gated by review completion."""

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core.config import Settings
from app.core.db import Base
from app.main import create_app


@pytest.fixture
def client():
    app = create_app(
        Settings(
            enabled_modules=(
                "health,rulepacks,auth,ingestion,findings,risk,review,drafting,export"
            ),
            database_url="sqlite:///:memory:",
        )
    )
    Base.metadata.create_all(app.state.ctx.registry.require("db.engine"))
    return TestClient(app)


def _auth(client):
    client.post(
        "/api/auth/signup",
        json={"email": "ex@x.com", "password": "hunter2hunter2", "org_name": "Acme"},
    )
    tok = client.post(
        "/api/auth/login", json={"email": "ex@x.com", "password": "hunter2hunter2"}
    ).json()["access_token"]
    return {"authorization": f"Bearer {tok}"}


def _opp_with_findings(client, headers):
    opp_id = client.post(
        "/api/ingestion/opportunities", json={"title": "Bridge"}, headers=headers
    ).json()["id"]
    client.post(
        f"/api/ingestion/opportunities/{opp_id}/documents",
        json={"filename": "gcc.pdf", "sample_text": "[p1]\nClause 5 — Scope. Build a bridge."},
        headers=headers,
    )
    client.post(f"/api/risk/opportunities/{opp_id}/run", headers=headers)
    return opp_id


def test_export_blocked_until_review_complete(client):
    headers = _auth(client)
    opp_id = _opp_with_findings(client, headers)
    r = client.get(
        f"/api/export/opportunities/{opp_id}?format=xlsx", headers=headers
    )
    assert r.status_code == 403
    assert r.json()["detail"] == "review_incomplete"


def test_export_xlsx_after_review(client):
    headers = _auth(client)
    opp_id = _opp_with_findings(client, headers)
    for f in client.get(f"/api/review/opportunities/{opp_id}/queue", headers=headers).json()[
        "findings"
    ]:
        client.post(
            f"/api/review/findings/{f['id']}", json={"decision": "accepted"}, headers=headers
        )

    r = client.get(
        f"/api/export/opportunities/{opp_id}?format=xlsx", headers=headers
    )
    assert r.status_code == 200
    assert r.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert "Bridge" in str(ws["A1"].value)
    assert ws["A4"].value == "Severity"


def test_export_bad_format(client):
    headers = _auth(client)
    opp_id = _opp_with_findings(client, headers)
    for f in client.get(f"/api/review/opportunities/{opp_id}/queue", headers=headers).json()[
        "findings"
    ]:
        client.post(
            f"/api/review/findings/{f['id']}", json={"decision": "accepted"}, headers=headers
        )

    r = client.get(
        f"/api/export/opportunities/{opp_id}?format=txt", headers=headers
    )
    assert r.status_code == 400
    assert r.json()["detail"] == "bad_format"
